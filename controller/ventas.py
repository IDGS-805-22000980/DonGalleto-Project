from calendar import monthrange
from flask import Blueprint, render_template, request, session, jsonify
from models.models import db,  Galleta, Venta, DetalleVenta
from datetime import datetime, date, timedelta
from flask_login import current_user, login_required
from sqlalchemy import func
from sqlalchemy.orm import joinedload
from flask import send_file
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

ventas = Blueprint('ventas', __name__)

@ventas.route('/indexVentas', methods=['GET'])
@login_required
def indexVentas():
    galletas = Galleta.query.all()
    return render_template('ventas/indexVentas.html', galletas=galletas)

######################## Metodos oara descontar del inventario ###############################

PRESENTACIONES = {
    'pieza': {'nombre': 'Por pieza', 'factor': 1, 'unidad': 'unidad'},
    'gramos': {'nombre': 'Por gramos', 'factor': 0.01, 'unidad': '100g'},
    '700g': {'nombre': 'Paquete 700g', 'factor': 7, 'unidad': 'paquete'},
    '1kg': {'nombre': 'Paquete 1kg', 'factor': 10, 'unidad': 'paquete'}
}

def calcular_galletas_a_descontar(presentacion, cantidad):
    """Calcula cuántas galletas se deben descontar del inventario"""
    if presentacion not in PRESENTACIONES:
        return None
        
    presentacion_data = PRESENTACIONES[presentacion]
    
    if presentacion == 'gramos':
        if cantidad % 100 != 0:  # Asegurar múltiplos de 100g
            return None
        return int(cantidad * presentacion_data['factor'])
    
    return int(cantidad * presentacion_data['factor'])

def reconstruir_presentacion(cantidad, galletas_descontadas):
    """Reconstruye el nombre de la presentación para mostrar al usuario"""
    if galletas_descontadas == cantidad:
        return f"{cantidad} {PRESENTACIONES['pieza']['unidad']}{'s' if cantidad > 1 else ''}"
    
    for key, data in PRESENTACIONES.items():
        if key != 'pieza' and galletas_descontadas == cantidad * data['factor']:
            if cantidad == 1:
                return f"1 {data['nombre']}"
            return f"{cantidad} {data['nombre']}{'s' if cantidad > 1 else ''}"
    
    # Si no coincide con ninguna presentación estándar
    return f"{galletas_descontadas} unidades equivalentes"

def determinar_presentacion(cantidad, factor):
    """Determina el nombre de la presentación basado en el factor"""
    for key, data in PRESENTACIONES.items():
        if data['factor'] == factor:
            if cantidad == 1:
                return data['nombre']
            return f"{cantidad} {data['nombre']}{'s' if cantidad > 1 else ''}"
    return "gramos"


######################################## Funciones para realizar una venta ########################################


 ##### Carga del carrito #####

@ventas.route('/carrito', methods=['GET'])
@login_required
def carrito():
    if 'carrito' not in session:
        return jsonify([])  # Si no hay carrito, devuelve una lista vacía
    return jsonify(session['carrito'])



###### Agregar producto al carrito #####

@ventas.route('/agregar_producto', methods=['POST'])
@login_required
def agregar_producto():
    try:
        if request.content_type != 'application/json':
            return jsonify({'error': 'Content-Type debe ser application/json'}), 400

        data = request.get_json()
        if not data:
            return jsonify({'error': 'No se recibieron datos'}), 400

        # Validaciones básicas
        nombre_galleta = data.get('nombre_galleta')
        presentacion = data.get('presentacion')
        cantidad = data.get('cantidad')

        if not all([nombre_galleta, presentacion, cantidad]):
            return jsonify({'error': 'Datos incompletos'}), 400

        # Convertir y validar cantidad
        try:
            cantidad = int(cantidad)
            if cantidad <= 0:
                return jsonify({'error': 'Cantidad debe ser positiva'}), 400
        except ValueError:
            return jsonify({'error': 'Cantidad inválida'}), 400

        # Validar presentación
        if presentacion not in PRESENTACIONES:
            return jsonify({'error': 'Presentación no válida'}), 400

        # Buscar galleta
        galleta = Galleta.query.filter_by(nombre=nombre_galleta).first()
        if not galleta:
            return jsonify({'error': 'Galleta no encontrada'}), 400

        # Calcular galletas a descontar
        galletas_a_descontar = calcular_galletas_a_descontar(presentacion, cantidad)
        if galletas_a_descontar is None:
            return jsonify({'error': 'Para gramos, use múltiplos de 100'}), 400

        # Verificar stock
        stock_actual = sum(inv.stock for inv in galleta.inventario if inv.disponible)
        if galletas_a_descontar > stock_actual:
            return jsonify({'error': f'Stock insuficiente. Disponible: {stock_actual}'}), 400

        # Calcular subtotal como float
        subtotal = float(galletas_a_descontar * galleta.precio)

        # Inicializar carrito si no existe
        if 'carrito' not in session:
            session['carrito'] = []

        # Buscar si el producto ya está en el carrito
        producto_existente = None
        for item in session['carrito']:
            if item['nombre_galleta'] == nombre_galleta and item['presentacion'] == presentacion:
                producto_existente = item
                break

        if producto_existente:
            # Actualizar producto existente
            producto_existente['cantidad'] += cantidad
            producto_existente['galletas_a_descontar'] += galletas_a_descontar
            producto_existente['subtotal'] += subtotal
        else:
            # Agregar nuevo producto
            session['carrito'].append({
                'nombre_galleta': nombre_galleta,
                'presentacion': presentacion,
                'cantidad': cantidad,
                'galletas_a_descontar': galletas_a_descontar,
                'precio_unitario': float(galleta.precio),
                'subtotal': subtotal,
                'presentacion_nombre': PRESENTACIONES[presentacion]['nombre']
            })

        session.modified = True

        return jsonify({
            'message': 'Producto agregado',
            'carrito': session['carrito']
        }), 200

    except Exception as e:
        print(f"❌ Error en agregar_producto: {str(e)}")
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    

####### Eliminar producto del carrito #######

@ventas.route('/eliminar_producto', methods=['POST'])
@login_required
def eliminar_producto():
    try:
        if 'carrito' not in session or not session['carrito']:
            return jsonify({'error': 'El carrito está vacío'}), 400

        data = request.get_json()
        if not data:
            return jsonify({'error': 'No se recibieron datos'}), 400

        nombre_galleta = data.get('nombre_galleta')
        presentacion = data.get('presentacion')

        if not nombre_galleta or not presentacion:
            return jsonify({'error': 'Datos incompletos'}), 400

        # Filtrar los productos para eliminar el que coincida
        nuevo_carrito = [
            item for item in session['carrito']
            if not (item['nombre_galleta'] == nombre_galleta and item['presentacion'] == presentacion)
        ]

        session['carrito'] = nuevo_carrito
        session.modified = True

        return jsonify({'message': 'Producto eliminado', 'carrito': session['carrito']}), 200

    except Exception as e:
        print(f"❌ Error en eliminar_producto: {e}")
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    

######### Guardar venta #########

@ventas.route('/guardar_venta', methods=['POST'])
@login_required
def guardar_venta():
    try:
        # Verificar carrito
        if 'carrito' not in session or not session['carrito']:
            return jsonify({'error': 'El carrito está vacío'}), 400

        # Obtener y validar datos
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No se recibieron datos'}), 400

        # Validar monto recibido
        try:
            monto_recibido = float(data.get('monto_recibido', 0))
            if monto_recibido <= 0:
                return jsonify({'error': 'Monto recibido inválido'}), 400
        except (TypeError, ValueError):
            return jsonify({'error': 'Monto debe ser numérico'}), 400

        # Calcular total de la venta
        total_venta = sum(float(item['subtotal']) for item in session['carrito'])
        total_venta = round(total_venta, 2)  # Redondear a 2 decimales

        # Validar monto suficiente
        if monto_recibido < total_venta:
            return jsonify({
                'error': f'Monto insuficiente. Total: ${total_venta:.2f}',
                'required': total_venta
            }), 400

        # Calcular cambio
        cambio = round(monto_recibido - total_venta, 2)

        # Calcular costos y ganancias
        costo_total = 0
        for item in session['carrito']:
            galleta = Galleta.query.filter_by(nombre=item['nombre_galleta']).first()
            if not galleta:
                return jsonify({'error': f'Galleta {item["nombre_galleta"]} no encontrada'}), 400
            costo_total += float(galleta.costo_galleta) * item['galletas_a_descontar']
        
        costo_total = round(costo_total, 2)
        ganancia = round(total_venta - costo_total, 2)

        # Crear nueva venta
        nueva_venta = Venta(
            usuario_id=current_user.id,
            fecha_venta=datetime.now(),
            total=total_venta,
            costo_total=costo_total,
            ganancia=ganancia,
            monto_recibido=monto_recibido,
            cambio=cambio
        )
        db.session.add(nueva_venta)
        db.session.flush()  # Para obtener el ID de la venta

        # Procesar cada item del carrito
        for item in session['carrito']:
            galleta = Galleta.query.filter_by(nombre=item['nombre_galleta']).first()
            cantidad_a_descontar = int(item['galletas_a_descontar'])

            # Verificar stock suficiente
            stock_total = sum(inv.stock for inv in galleta.inventario if inv.disponible)
            if cantidad_a_descontar > stock_total:
                db.session.rollback()
                return jsonify({'error': f'Stock insuficiente para {galleta.nombre}'}), 400

            # Descontar del inventario
            for inventario in galleta.inventario:
                if inventario.disponible and cantidad_a_descontar > 0:
                    if inventario.stock >= cantidad_a_descontar:
                        inventario.stock -= cantidad_a_descontar
                        cantidad_a_descontar = 0
                    else:
                        cantidad_a_descontar -= inventario.stock
                        inventario.stock = 0
                    
                    if inventario.stock == 0:
                        inventario.disponible = False

            # Crear detalle de venta
            detalle = DetalleVenta(
                venta_id=nueva_venta.id,
                galleta_id=galleta.id,
                cantidad=item['cantidad'],
                galletas_descontadas=item['galletas_a_descontar'],
                precio_unitario=float(item['precio_unitario']),
                subtotal=float(item['subtotal'])
            )
            db.session.add(detalle)

        # Confirmar cambios en la base de datos
        db.session.commit()
        
        # Limpiar carrito de la sesión
        session.pop('carrito', None)

        # Preparar respuesta
        response_data = {
            'message': 'Venta guardada con éxito',
            'cambio': cambio,
            'total': total_venta,
            'venta_id': nueva_venta.id
        }

        # Agregar URL de ticket si se solicitó imprimir
        if data.get('imprimir_ticket', False):
            response_data['ticket_url'] = f'/ventas/ticket/{nueva_venta.id}'

        return jsonify(response_data), 200

    except Exception as e:
        db.session.rollback()
        print(f"❌ Error en guardar_venta: {str(e)}")
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    



 ######## Generar ticket de venta #########

@ventas.route('/ticket/<int:venta_id>')
@login_required
def generar_ticket(venta_id):
    venta = Venta.query.get_or_404(venta_id)
    detalles = []
    
    for detalle in venta.detalles:
        galleta = Galleta.query.get(detalle.galleta_id)
        
        # Calcular el factor de presentación
        factor = detalle.galletas_descontadas / detalle.cantidad
        
        # Determinar la presentación
        presentacion = determinar_presentacion(detalle.cantidad, factor)
        
        detalles.append({
            'nombre': galleta.nombre,
            'presentacion': presentacion,
            'cantidad': detalle.cantidad,
            'precio': float(detalle.precio_unitario),
            'subtotal': float(detalle.subtotal)
        })
    
    return render_template('ventas/ticket.html',
                         venta=venta,
                         detalles=detalles,
                         fecha=venta.fecha_venta.strftime('%d/%m/%Y %H:%M:%S')) 


    
################################# Listar ventas #################################

@ventas.route('/listaVentas', methods=['GET'])
def listaVentas():
    return render_template('ventas/listaVentas.html')


@ventas.route('/listar_ventas', methods=['GET'])
@login_required
def listar_ventas():
    try:
        fecha = request.args.get('fecha')
        
        if fecha:
            try:
                fecha = datetime.strptime(fecha, "%Y-%m-%d").date()
                ventas = Venta.query.filter(db.func.date(Venta.fecha_venta) == fecha)\
                                  .order_by(Venta.fecha_venta.desc()).all()
            except ValueError:
                return jsonify({'error': 'Formato de fecha inválido, use YYYY-MM-DD'}), 400
        else:
            ventas = Venta.query.order_by(Venta.fecha_venta.desc()).all()

        ventas_data = []
        for venta in ventas:
            ventas_data.append({
                'id': venta.id,
                'fecha': venta.fecha_venta.strftime('%Y-%m-%d %H:%M:%S'),
                'total': float(venta.total),
                'ganancia': float(venta.ganancia),
                'monto_recibido': float(venta.monto_recibido),
                'cambio': float(venta.cambio)
            })

        return jsonify(ventas_data), 200

    except Exception as e:
        print(f"❌ Error en listar_ventas: {e}")
        return jsonify({'error': 'Error al obtener el listado de ventas'}), 500
    

@ventas.route('/detalles_venta/<int:id_venta>')
@login_required
def detalles_venta(id_venta):
    try:
        venta = Venta.query.options(
            joinedload(Venta.detalles).joinedload(DetalleVenta.galleta)
        ).get_or_404(id_venta)
        
        detalles = []
        for detalle in venta.detalles:
            # Calcular el factor de presentación
            factor = detalle.galletas_descontadas / detalle.cantidad if detalle.cantidad > 0 else 0
            
            detalles.append({
                'galleta': detalle.galleta.nombre,
                'presentacion': determinar_presentacion(detalle.cantidad, factor),
                'cantidad': detalle.cantidad,
                'precio_unitario': float(detalle.precio_unitario),
                'subtotal': float(detalle.subtotal),
                'galletas_descontadas': detalle.galletas_descontadas
            })
        
        return jsonify({
            'id': venta.id,
            'fecha': venta.fecha_venta.strftime('%Y-%m-%d %H:%M:%S'),
            'total': float(venta.total),
            'ganancia': float(venta.ganancia),
            'monto_recibido': float(venta.monto_recibido),
            'cambio': float(venta.cambio),
            'detalles': detalles
        })

    except Exception as e:
        print(f"Error al obtener detalles de venta: {str(e)}")
        return jsonify({'error': 'No se pudieron obtener los detalles de la venta'}), 500


    

################################# Generar corte #################################

@ventas.route('/generar_corte', methods=['GET'])
@login_required
def generar_corte():
    try:
        tipo_corte = request.args.get('tipo', 'dia')
        fecha_str = request.args.get('fecha')
        
        if fecha_str:
            try:
                fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except ValueError:
                return jsonify({'error': 'Formato de fecha inválido, use YYYY-MM-DD'}), 400
        else:
            fecha = datetime.now().date()

        # Determinar rango de fechas
        if tipo_corte == 'dia':
            inicio = fecha
            fin = fecha
        elif tipo_corte == 'semana':
            inicio = fecha - timedelta(days=fecha.weekday())
            fin = inicio + timedelta(days=6)
        elif tipo_corte == 'mes':
            inicio = fecha.replace(day=1)
            fin = fecha.replace(day=monthrange(fecha.year, fecha.month)[1])
        else:
            return jsonify({'error': 'Tipo de corte inválido'}), 400

        # Obtener ventas en el rango
        ventas_corte = Venta.query.filter(
            db.func.date(Venta.fecha_venta) >= inicio,
            db.func.date(Venta.fecha_venta) <= fin
        ).all()

        # Calcular métricas principales
        total_ventas = sum(float(v.total) for v in ventas_corte)
        total_costos = sum(float(v.costo_total) for v in ventas_corte)
        total_ganancias = sum(float(v.ganancia) for v in ventas_corte)
        num_ventas = len(ventas_corte)

        # Presentaciones más vendidas (las que más se repiten)
        presentaciones_data = db.session.query(
            (DetalleVenta.galletas_descontadas / DetalleVenta.cantidad).label('tipo_presentacion'),
            func.count().label('total_repeticiones'),
            func.sum(DetalleVenta.cantidad).label('total_cantidad')
        ).join(Venta).filter(
            func.date(Venta.fecha_venta) >= inicio,
            func.date(Venta.fecha_venta) <= fin
        ).group_by('tipo_presentacion').order_by(func.count().desc()).limit(5)

        presentaciones_mas_vendidas = []
        for tipo, repeticiones, cantidad in presentaciones_data:
            factor = float(tipo)
            presentaciones_mas_vendidas.append({
                'presentacion': determinar_presentacion(1, factor),
                'repeticiones': int(repeticiones),
                'cantidad_total': int(cantidad)
            })

        # Galletas más vendidas (por cantidad descontada)
        galletas_data = db.session.query(
            Galleta.nombre,
            func.sum(DetalleVenta.galletas_descontadas).label('total_descontado'),
            func.sum(DetalleVenta.subtotal).label('total_venta')
        ).join(DetalleVenta).join(Venta).filter(
            func.date(Venta.fecha_venta) >= inicio,
            func.date(Venta.fecha_venta) <= fin
        ).group_by(Galleta.nombre).order_by(func.sum(DetalleVenta.galletas_descontadas).desc()).limit(5)

        galletas_mas_vendidas = []
        for nombre, total_descontado, total_venta in galletas_data:
            galletas_mas_vendidas.append({
                'nombre': nombre,
                'total_descontado': int(total_descontado),
                'total_venta': float(total_venta)
            })

        # Datos para gráficas
        datos_graficas = {
            'presentaciones': {
                'labels': [p['presentacion'] for p in presentaciones_mas_vendidas],
                'data': [p['repeticiones'] for p in presentaciones_mas_vendidas],
                'cantidades': [p['cantidad_total'] for p in presentaciones_mas_vendidas]
            },
            'galletas': {
                'labels': [g['nombre'] for g in galletas_mas_vendidas],
                'data': [g['total_descontado'] for g in galletas_mas_vendidas],
                'ventas': [g['total_venta'] for g in galletas_mas_vendidas]
            }
        }

        # Detalle de ventas
        detalle_ventas = [{
            'id': v.id,
            'fecha': v.fecha_venta.strftime('%Y-%m-%d %H:%M:%S'),
            'total': float(v.total),
            'ganancia': float(v.ganancia)
        } for v in ventas_corte]

        return jsonify({
            'tipo_corte': tipo_corte,
            'fecha_inicio': inicio.strftime('%Y-%m-%d'),
            'fecha_fin': fin.strftime('%Y-%m-%d'),
            'total_ventas': total_ventas,
            'total_costos': total_costos,
            'total_ganancias': total_ganancias,
            'num_ventas': num_ventas,
            'presentaciones_mas_vendidas': presentaciones_mas_vendidas,
            'galletas_mas_vendidas': galletas_mas_vendidas,
            'datos_graficas': datos_graficas,
            'detalle_ventas': detalle_ventas
        })

    except Exception as e:
        print(f"Error en generar_corte: {str(e)}")
        return jsonify({'error': 'Error al generar el corte'}), 500
    


 ################ exportar el excel ##############   

@ventas.route('/exportar_corte_excel', methods=['POST'])
@login_required
def exportar_corte_excel():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No se recibieron datos'}), 400

        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Corte de Ventas"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        # Encabezado
        ws.append(["Corte de Ventas"])
        ws.merge_cells('A1:D1')
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")

        # Fechas
        ws.append(["Desde:", data['fecha_inicio'], "Hasta:", data['fecha_fin']])
        ws.append(["Total Ventas:", data['total_ventas'], "Total Ganancias:", data['total_ganancias']])
        ws.append(["Número de Ventas:", data['num_ventas'], "Margen:", f"{(data['total_ganancias'] / data['total_ventas'] * 100 if data['total_ventas'] > 0 else 0):.2f}%"])

        # Presentaciones más vendidas
        ws.append([])
        ws.append(["Presentaciones Más Vendidas", "Repeticiones", "Cantidad Total"])
        for p in data['presentaciones_mas_vendidas']:
            ws.append([p['presentacion'], p['repeticiones'], p['cantidad_total']])

        # Galletas más vendidas
        ws.append([])
        ws.append(["Galletas Más Vendidas", "Cantidad Descontada", "Total Ventas"])
        for g in data['galletas_mas_vendidas']:
            ws.append([g['nombre'], g['total_descontado'], g['total_venta']])

        # Detalle de ventas
        ws.append([])
        ws.append(["Detalle de Ventas"])
        ws.merge_cells(f'A{ws.max_row}:D{ws.max_row}')
        ws.append(["ID", "Fecha/Hora", "Total", "Ganancia"])
        for v in data['detalle_ventas']:
            ws.append([v['id'], v['fecha'], v['total'], v['ganancia']])

        # Aplicar formatos
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                if cell.row in [1, 2, 3, 4, 6, 8 + len(data['presentaciones_mas_vendidas']), 
                              10 + len(data['presentaciones_mas_vendidas']) + len(data['galletas_mas_vendidas'])]:
                    cell.font = Font(bold=True)
                if cell.row in [6, 8 + len(data['presentaciones_mas_vendidas']), 
                              10 + len(data['presentaciones_mas_vendidas']) + len(data['galletas_mas_vendidas'])]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

        # Ajustar anchos de columna
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Guardar en un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Enviar el archivo
        fecha_reporte = datetime.now().strftime("%Y-%m-%d")
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"corte_ventas_{fecha_reporte}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"Error al exportar corte a Excel: {str(e)}")
        return jsonify({'error': 'Error al generar el reporte Excel'}), 500
    


@ventas.route('/corte')
@login_required
def corte():
    hoy = datetime.now().date().strftime('%Y-%m-%d')
    return render_template('ventas/corte.html', hoy=hoy)